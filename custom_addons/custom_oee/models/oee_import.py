# -*- coding: utf-8 -*-
# Import file Excel rekap OEE (sheet "Daily Entry") menjadi record oee.entry.
# Setiap kali import dibuat satu batch, sehingga user bisa menghapus
# hasil import tertentu tanpa mengganggu data input manual / import lain.
import base64
import io
from datetime import datetime

from odoo import api, fields, models
from odoo.exceptions import UserError


class OeeImportBatch(models.Model):
    _name = 'oee.import.batch'
    _description = 'Batch Import Excel OEE'
    _order = 'create_date desc'

    name = fields.Char(string='Nama File', required=True)
    jumlah_baris = fields.Integer(string='Jumlah Baris')
    periode = fields.Char(string='Periode Data')
    entry_ids = fields.One2many('oee.entry', 'import_batch_id', string='Data OEE')

    # ==== Helper untuk aplikasi OWL menu Import ====
    @api.model
    def oee_import_daftar(self):
        """Daftar semua batch import untuk ditampilkan di halaman Import."""
        hasil = []
        for batch in self.search([]):
            hasil.append({
                'id': batch.id,
                'name': batch.name,
                'tanggal': fields.Datetime.context_timestamp(
                    batch, batch.create_date).strftime('%d/%m/%Y %H:%M'),
                'jumlah': batch.jumlah_baris,
                'periode': batch.periode or '-',
            })
        return hasil

    @api.model
    def oee_import_hapus(self, batch_id):
        """Hapus satu batch beserta seluruh baris data hasil import-nya."""
        batch = self.browse(int(batch_id)).exists()
        if batch:
            info = "Hapus import '%s' (%s baris, %s)" % (
                batch.name, batch.jumlah_baris, batch.periode or '-')
            # riwayat cukup 1 baris untuk seluruh batch, bukan per record
            batch.entry_ids.with_context(skip_oee_riwayat=True).unlink()
            batch.unlink()
            self.env['oee.riwayat']._catat('hapus_import', info)
        return True

    @api.model
    def oee_import_excel(self, filename, file_b64):
        """Baca sheet 'Daily Entry' dari file Excel lalu buat record oee.entry.

        Kolom yang dibaca (mengikuti struktur file rekap):
        A=Tanggal, B=Shift, C=No Mesin, D=Mold, E=Down Time (menit),
        G=Total waktu tersedia (menit), I=HPR Actual (pcs),
        J=Target Production (detik) -> ideal CT = J / I, N=Total Reject (pcs).
        """
        try:
            import openpyxl
        except ImportError:
            raise UserError('Library openpyxl tidak tersedia di server.')

        try:
            data = base64.b64decode(file_b64)
            wb = openpyxl.load_workbook(
                io.BytesIO(data), data_only=True, read_only=True)
        except Exception:
            raise UserError(
                'File tidak bisa dibaca. Pastikan file Excel (.xlsx) yang benar.')

        if 'Daily Entry' in wb.sheetnames:
            ws = wb['Daily Entry']
        else:
            # fallback: pakai sheet pertama jika nama sheet berbeda
            ws = wb.worksheets[0]

        def angka(v):
            # konversi aman ke float (formula yang tidak ter-cache jadi 0)
            try:
                return float(v)
            except (TypeError, ValueError):
                return 0.0

        vals_list = []
        tgl_min = tgl_max = None
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or len(row) < 14:
                continue
            tanggal, shift, no_mesin, mold = row[0], row[1], row[2], row[3]
            if not isinstance(tanggal, datetime) or not mold:
                continue

            shift_str = str(int(angka(shift)))
            if shift_str not in ('1', '2', '3'):
                shift_str = '1'

            hpr = int(angka(row[8]))
            target_sec = angka(row[9])
            vals_list.append({
                'tanggal': tanggal.date(),
                'shift': shift_str,
                'no_mesin': str(no_mesin) if no_mesin is not None else '-',
                'mold': str(mold).strip(),
                'downtime_unplanned': angka(row[4]),
                'total_menit': angka(row[6]) or 720.0,
                'counter_awal': 0,
                'counter_akhir': hpr,
                'ideal_ct': round(target_sec / hpr, 2) if hpr else 0.0,
                'total_reject': int(angka(row[13])),
            })
            tgl = tanggal.date()
            tgl_min = tgl if not tgl_min or tgl < tgl_min else tgl_min
            tgl_max = tgl if not tgl_max or tgl > tgl_max else tgl_max

        wb.close()
        if not vals_list:
            raise UserError(
                'Tidak ada baris data yang bisa dibaca dari file ini. '
                'Pastikan formatnya sama dengan file rekap OEE (sheet "Daily Entry").')

        batch = self.create({
            'name': filename,
            'jumlah_baris': len(vals_list),
            'periode': '%s s/d %s' % (
                tgl_min.strftime('%d/%m/%Y'), tgl_max.strftime('%d/%m/%Y')),
        })
        for v in vals_list:
            v['import_batch_id'] = batch.id

        # create per potongan agar tidak membebani memori untuk file besar;
        # riwayat per record dilewati — dicatat 1 baris untuk seluruh import
        Entry = self.env['oee.entry'].with_context(skip_oee_riwayat=True)
        for i in range(0, len(vals_list), 2000):
            Entry.create(vals_list[i:i + 2000])

        self.env['oee.riwayat']._catat(
            'import', "Import Excel '%s' (%s baris, %s)" % (
                filename, len(vals_list), batch.periode))
        return {'id': batch.id, 'jumlah': len(vals_list)}
